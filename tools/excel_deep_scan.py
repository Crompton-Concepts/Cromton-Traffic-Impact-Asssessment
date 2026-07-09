"""Deep scan for VCR, hourly, and peak labels in Excel templates."""
import os
import re
import openpyxl

FOLDER = os.path.join(os.path.dirname(__file__), '..', 'excel calculation templates')

PATTERNS = {
    'vcr': re.compile(r'vcr|v/c|vol.*cap|capacity|design vol|/(\$?[A-Z]+\$?\d+)', re.I),
    'peak': re.compile(r'12%|15%|60%|85%|peak|aadt|hrly|hourly|design volume', re.I),
    'queue': re.compile(r'queue|ROUNDUP.*\*|2\.4|single lane|SLC|merge|taper', re.I),
}


def deep_scan(path, max_row=200, max_col=100):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    hits = {k: [] for k in PATTERNS}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                v = cell.value
                if not v:
                    continue
                s = str(v)
                coord = f'{sn}!{cell.coordinate}'
                for key, rx in PATTERNS.items():
                    if rx.search(s) and len(hits[key]) < 15:
                        hits[key].append(f'{coord}: {s[:140]}')
    wb.close()
    return hits


if __name__ == '__main__':
    for fname in sorted(os.listdir(FOLDER)):
        if not fname.endswith(('.xlsx', '.xlsm')):
            continue
        print(f'\n{"="*60}\n{fname}\n{"="*60}')
        h = deep_scan(os.path.join(FOLDER, fname))
        for k, items in h.items():
            print(f'\n--- {k.upper()} ({len(items)}) ---')
            for i in items:
                print(f'  {i}')
