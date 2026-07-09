"""Compare queue/VCR formulas across Excel calculation templates."""
import openpyxl
import os
import re

FOLDER = os.path.join(os.path.dirname(__file__), '..', 'excel calculation templates')


def get_queue_factors(path):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=80, max_col=40):
            for cell in row:
                if cell.value == 'Queue Length factors':
                    r = cell.row
                    factors = {}
                    for dr in range(1, 8):
                        tcell = ws.cell(r + dr, cell.column)
                        zcell = ws.cell(r + dr, cell.column + 1)
                        aacell = ws.cell(r + dr, cell.column + 2)
                        try:
                            m = int(float(tcell.value))
                            factors[m] = (zcell.value, aacell.value)
                        except (TypeError, ValueError):
                            pass
                    out[sn] = factors
    wb.close()
    return out


def find_pattern(path, pattern, max_hits=8):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    refs = []
    rx = re.compile(pattern, re.I)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=80):
            for cell in row:
                v = cell.value
                if v and isinstance(v, str) and rx.search(v):
                    refs.append(f'{sn}!{cell.coordinate}: {v[:100]}')
                    if len(refs) >= max_hits:
                        wb.close()
                        return refs
    wb.close()
    return refs


def main():
    for fname in sorted(os.listdir(FOLDER)):
        if not fname.endswith(('.xlsx', '.xlsm')):
            continue
        path = os.path.join(FOLDER, fname)
        print(f'=== {fname} ===')
        qf = get_queue_factors(path)
        for sn, factors in qf.items():
            print(f'  Sheet [{sn}] queue factors (min -> LV mult, HV mult):')
            for m in sorted(factors):
                lv, hv = factors[m]
                print(f'    {m:2d} min: LV={lv}, HV={hv}')
        refs85 = find_pattern(path, r'85%|0\.85|\*85')
        print(f'  85% factor formulas ({len(refs85)}):')
        for r in refs85:
            print(f'    {r}')
        slc = find_pattern(path, r'Single Lane.*=')
        print(f'  Single Lane Closure formulas ({len(slc)}):')
        for r in slc:
            print(f'    {r}')
        merge = find_pattern(path, r'merge|taper|1\.5|2\.0|contingency', max_hits=4)
        print(f'  Merge/taper refs ({len(merge)}):')
        for r in merge:
            print(f'    {r}')
        print()


if __name__ == '__main__':
    main()
