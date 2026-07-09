"""Extract council-specific formulas from Excel calculation templates."""
import json
import os
import re
from collections import defaultdict

import openpyxl

FOLDER = os.path.join(os.path.dirname(__file__), '..', 'excel calculation templates')
OUT = os.path.join(os.path.dirname(__file__), 'excel_council_audit.md')

PEAK_PATTERNS = [
    (r'12\s*%\s*\*\s*85%|12%.*85%|\*12%\*85%|\(.*?12%.*?\)\*85%', '12% × 85% of AADT'),
    (r'15\s*%\s*\*\s*85%|15%.*85%|\*15%\)\*85%|\(.*?15%.*?\)\*85%', '15% × 85% of AADT'),
    (r'60\s*%\s*\*\s*85%|60%.*85%|\*60%\*85%', '60% × 85% of period count / lanes'),
    (r'85\s*%\s*of\s*12\s*%|85% of 12%', '85% of 12% of AADT (design volume label)'),
    (r'100\*.*\/42.*85%|\(100\*.*\/42\)\*85%', '5-min count scaled to hour × 85%'),
    (r'12\s*%\s*of\s*AADT|AADT.*12%|12%\s*AADT', '12% of AADT (no 85%)'),
    (r'15\s*%\s*of\s*AADT|AADT.*15%|15%\s*AADT', '15% of AADT (no 85%)'),
    (r'60\s*%\s*peak|peak.*60%|\*60%', '60% peak period factor on counts'),
    (r'15\s*%\s*peak|peak.*15%|\*15%', '15% peak period factor on counts'),
    (r'measured|hourly\s*profile|count\s*sheet|traffic count', 'Measured hourly from count profile'),
]

VCR_PATTERNS = [
    (r'12\s*%\s*AADT|AADT\s*\*\s*12%|\*12%', 'Design volume = 12% of AADT'),
    (r'15\s*%\s*AADT|AADT\s*\*\s*15%|\*15%', 'Design volume = 15% of AADT'),
    (r'design\s*volume|design\s*cap|DV\b', 'Austroads design capacity (DV) denominator'),
    (r'SUM\(.*\)\*\$?[A-Z]+\$?\d+\/\$', 'SLC VCR: SUM(hourly) × lanes / capacity'),
    (r'\/\$?[A-Z]+\$\d+.*VCR|VCR.*\/', 'Hourly demand / capacity ratio'),
    (r'Single\s*Lane|SLC|lane\s*closure', 'Single lane closure scaling'),
]


def get_queue_factors(ws, sn):
    factors = {}
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=40):
        for cell in row:
            if cell.value == 'Queue Length factors':
                r, c = cell.row, cell.column
                for dr in range(1, 10):
                    tcell = ws.cell(r + dr, c)
                    zcell = ws.cell(r + dr, c + 1)
                    aacell = ws.cell(r + dr, c + 2)
                    try:
                        m = int(float(tcell.value))
                        factors[m] = (zcell.value, aacell.value)
                    except (TypeError, ValueError):
                        pass
                return factors
    return factors


def scan_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    info = {
        'sheets': wb.sheetnames,
        'queue_factors': {},
        'formulas_85': [],
        'formulas_peak': [],
        'formulas_vcr': [],
        'formulas_hourly': [],
        'formulas_queue_calc': [],
        'labels': [],
        'title_cells': [],
    }

    for sn in wb.sheetnames:
        ws = wb[sn]
        qf = get_queue_factors(ws, sn)
        if qf:
            info['queue_factors'][sn] = qf

        for row in ws.iter_rows(min_row=1, max_row=150, max_col=80):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                coord = f'{sn}!{cell.coordinate}'
                if not isinstance(v, str):
                    continue

                if v == 'Queue Length factors':
                    continue  # handled above

                if v.startswith('='):
                    f = v
                    if re.search(r'85%|0\.85|\*85\b', f, re.I):
                        info['formulas_85'].append(f'{coord}: {f}')
                    if re.search(r'12%|15%|60%|AADT|ROUNDUP.*%|peak|Hrly|hourly', f, re.I):
                        info['formulas_peak'].append(f'{coord}: {f}')
                        info['formulas_hourly'].append(f'{coord}: {f}')
                    if re.search(r'capacity|VCR|/\$?[A-Z]+\$?\d+|design|DV|Austroads|SUM\(', f, re.I):
                        info['formulas_vcr'].append(f'{coord}: {f}')
                    if re.search(r'ROUNDUP\s*\(\s*[A-Z]+\d+\s*\*\s*[A-Z]+\d+|queue|2\.4|\*6\b|\*20\b', f, re.I):
                        info['formulas_queue_calc'].append(f'{coord}: {f}')
                elif len(v) > 8 and re.search(r'traffic|analysis|tia|hospital|impact|queue|vcr|aadt|peak|design', v, re.I):
                    info['title_cells'].append(f'{coord}: {v[:100]}')
                elif re.search(r'85%|12%|15%|60%|peak hour|design volume|k-factor|k factor|aadt|vcr|queue length', v, re.I):
                    info['labels'].append(f'{coord}: {v[:120]}')

    wb.close()
    return info


def infer_peak_formula(info, fname):
    hits = []
    corpus = '\n'.join(info['formulas_peak'] + info['formulas_85'] + info['labels'])
    for pat, label in PEAK_PATTERNS:
        if re.search(pat, corpus, re.I):
            if label not in hits:
                hits.append(label)

    if 'Brisbane' in fname and not any('85%' in h or '12%' in h or '15%' in h for h in hits):
        hits.insert(0, 'Measured peak hour from Traffic Count sheets (no AADT × peak% formula)')
    if 'Logan' in fname and not hits:
        hits.append('Measured hourly from count data (Logan TIA macro workbook)')

    if not hits:
        pcts = sorted(set(re.findall(r'(\d+)%', corpus)), key=int)
        if pcts:
            hits.append(f'Formula uses percentage factors: {", ".join(pcts + ["%"])}')
        else:
            hits.append('Unknown — inspect count input sheets')
    return hits


def uses_k_factor(info):
    corpus = '\n'.join(info['formulas_85'] + info['labels'])
    if re.search(r'85%|0\.85|\*85\b', corpus, re.I):
        return True, info['formulas_85'][:8]
    return False, []


def infer_vcr(info):
    hits = []
    corpus = '\n'.join(info['formulas_vcr'] + info['labels'])
    for pat, label in VCR_PATTERNS:
        if re.search(pat, corpus, re.I) and label not in hits:
            hits.append(label)
    samples = info['formulas_vcr'][:10]
    if not hits:
        hits = ['Hourly directional volume / Austroads design capacity per lane']
    return hits, samples


def infer_hourly(info, fname):
    corpus = '\n'.join(info['formulas_hourly'] + info['formulas_peak'] + info['formulas_85'])
    notes = []

    if re.search(r'ROUNDUP\([^)]*60%\*85%[^)]*/(\$?[A-Z]+\$?\d+|\d+)', corpus, re.I):
        notes.append('TMR-style: `ROUNDUP(period_total × 60% × 85% / lanes)` for AM/PM')
    if re.search(r'ROUNDUP\([^)]*15%\*85%[^)]*/(\$?[A-Z]+\$?\d+|\d+)', corpus, re.I):
        notes.append('Off-peak: `ROUNDUP(period_total × 15% × 85% / lanes)`')
    if re.search(r'ROUNDUP\(\([^)]*15%\)\*85%', corpus, re.I):
        notes.append('AADT path: `ROUNDUP(AADT × 15% × 85%)` total peak; HV split via HV% cell')
    if re.search(r'ROUNDUP\(\([^)]*12%\)\*85%|12%.*85%', corpus, re.I):
        notes.append('AADT path: `ROUNDUP(AADT × 12% × 85%)` (5-min template design check)')
    if re.search(r'\(100\*[^)]*\/42\)\*85%|100\*.*\/42.*85%', corpus, re.I):
        notes.append('5-min counts: `ROUNDUP((100 × 5min_count / 42) × 85%)` hourly equivalent')
    if re.search(r'\/(\$?[A-Z]+\$?\d+|\d+)\s*,\s*0\)', corpus, re.I):
        notes.append('Per-lane hourly: divide by lane count after peak/K factors')

    if 'Brisbane' in fname or 'Logan' in fname:
        notes.insert(0, 'Primary: measured hourly avg/lane from traffic count profile (not AADT-derived)')

    if not notes:
        notes.append('Direct from measured hourly profile / count sheets')

    return notes, info['formulas_hourly'][:8]


def council_from_filename(fname):
    mapping = {
        'Traffic ANALYSIS_Brisbane.xlsx': 'Brisbane City Council (BCC)',
        'Traffic ANALYSIS_TMR.xlsx': 'TMR (Queensland Dept of Transport)',
        'Traffic Analysis_5 minute counts.xlsx': 'Generic — 5-minute count methodology',
        'Traffic Analysis_AADT only.xlsx': 'Generic — AADT-only (no count profile)',
        'Traffic Analysis_GCCC Counts.xlsx': 'Gold Coast City Council (GCCC)',
        'Traffic Analysis_Ipswich City Council.xlsx': 'Ipswich City Council',
        'Traffic Analysis_Logan.xlsm': 'Logan City Council',
        'Traffic Analysis_Toowoomba Counts.xlsx': 'Toowoomba Regional Council',
    }
    return mapping.get(fname, fname)


def quirks(fname, info):
    q = []
    if 'Brisbane' in fname:
        q.append('Multi-site workbook (Bowen Tce + Traffic Count sheets); peak from measured counts')
        q.append('No 0.85 K-factor in formulas — uses raw count profile')
        q.append('Traffic Count sheet notes reference Gateway merge location (300 m south)')
    if 'TMR' in fname:
        q.append('Period totals (AM/OP/PM/EV) with AM/PM = 60%, OP/EV = 15% peak factors')
        q.append('All hourly avg/lane cells use explicit `× 85%` K-factor')
        q.append('Example: `=ROUNDUP(AL17*60%*85%/$AL$10,0)` (LV AM hourly per lane)')
    if '5 minute' in fname:
        q.append('Design volume label: "85% of 12% of AADT" for cross-check against counts')
        q.append('Hourly from 5-min interval: `ROUNDUP((100*count/42)*85%,0)` — factor 100/42 ≈ 2.381')
    if 'AADT only' in fname:
        q.append('No traffic count profile — entirely AADT-driven')
        q.append('Peak LV: `ROUNDUP(AADT*15%*85%,0) - ROUNDUP((AADT*15%*85%)*HV%,0)`')
        q.append('Peak HV: `ROUNDUP((AADT*15%*85%)*HV%,0)`')
    if 'GCCC' in fname:
        q.append('Two example sites: TIA - Hospital, Rosewood — same formula structure')
        q.append('Count-based primary input + 15%×85% AADT fallback on summary sheet')
    if 'Ipswich' in fname:
        q.append('Identical formula layout to AADT-only template (15%×85% + HV split)')
    if 'Logan' in fname:
        q.append('Macro-enabled workbook (.xlsm)')
        q.append('Sheet: "Traffic Impact and Queue Length" — count-driven, no 85% formulas in scan range')
    if 'Toowoomba' in fname:
        q.append('Hospital example site with GCCC-style 15%×85% formulas')
        q.append('Count sheets drive hourly profile; AADT formulas for design volume check')

    merged = {}
    for fac in info['queue_factors'].values():
        merged.update(fac)
    if 30 in merged:
        q.append('Includes 30-min queue duration row (LV×36, HV×120) — extended hold scenario')
    return q


def queue_summary(info):
    merged = {}
    for sn, fac in info['queue_factors'].items():
        merged.update(fac)
    if not merged:
        return {}, 'Not found in scanned sheets'
    lines = []
    for m in sorted(merged):
        lv, hv = merged[m]
        lines.append(f'{m} min: LV×{lv}, HV×{hv}')
    return merged, '; '.join(lines)


def render_md(audits):
    lines = [
        '# Excel Council Template Audit',
        '',
        'Structured comparison of all workbooks in `excel calculation templates/`.',
        'Generated by `tools/excel_council_audit.py` using openpyxl.',
        '',
        'Use alongside `tools/numerical_compare.py` and app logic in `app.js` / `calc/tia-calc.js`.',
        '',
        '## Summary Table',
        '',
        '| File | Council | Peak volume | 0.85 K | Queue (2/5/10/15 min) | Hourly avg | VCR basis |',
        '|------|---------|-------------|--------|------------------------|------------|-----------|',
    ]
    for a in audits:
        q = a['queue_factors']
        qshort = '/'.join(str(q.get(m, ('?', '?'))[0]) for m in (2, 5, 10, 15))
        lines.append(
            f"| {a['file']} | {a['council']} | {a['peak_primary']} | "
            f"{'Yes' if a['k_factor'] else 'No'} | LV mult: {qshort} | {a['hourly_primary']} | {a['vcr_primary']} |"
        )

    lines += ['', '---', '']
    for a in audits:
        lines += [
            f"## {a['file']}",
            '',
            f"**Council / source:** {a['council']}",
            f"**Sheets:** {', '.join(a['sheets'])}",
            '',
            '### 1. Peak volume formula',
            '',
        ]
        for p in a['peak_formulas']:
            lines.append(f'- {p}')
        if a['peak_samples']:
            lines += ['', 'Representative formulas:', '']
            for s in a['peak_samples']:
                lines.append(f'- `{s}`')

        lines += ['', '### 2. Queue length factors (LV / HV multipliers)', '', '| Duration | LV mult | HV mult | Notes |', '|----------|---------|---------|-------|']
        q = a['queue_factors']
        spacing_note = {2: '6 m × 0.4', 5: '6 m × 1 (base)', 10: '6 m × 2', 15: '6 m × 3', 30: '6 m × 6'}
        for m in (2, 5, 10, 15, 30):
            if m in q:
                lv, hv = q[m]
                lines.append(f'| {m} min | {lv} | {hv} | Equivalent spacing: {spacing_note.get(m, "—")} |')
        if not q:
            lines.append('| — | — | — | Not found |')

        lines += ['', 'Queue calc pattern: `ROUNDUP(LV_5min × LV_mult + HV_5min × HV_mult)` from hourly/5-min volumes.', '']

        lines += ['### 3. VCR / capacity formula', '']
        for v in a['vcr_notes']:
            lines.append(f'- {v}')
        if a['vcr_samples']:
            lines += ['', 'Sample formulas:', '']
            for s in a['vcr_samples']:
                lines.append(f'- `{s}`')

        lines += [
            '',
            f"### 4. 0.85 K-factor: **{'Yes' if a['k_factor'] else 'No'}**",
            '',
        ]
        if a['k_samples']:
            for s in a['k_samples']:
                lines.append(f'- `{s}`')
        else:
            lines.append('- No explicit 85% multiplier — volumes taken from measured counts at face value')

        lines += ['', '### 5. Hourly average calculation', '']
        for h in a['hourly_notes']:
            lines.append(f'- {h}')
        if a['hourly_samples']:
            lines += ['', 'Sample formulas:', '']
            for s in a['hourly_samples']:
                lines.append(f'- `{s}`')

        lines += ['', '### 6. Council-specific quirks', '']
        for qk in a['quirks']:
            lines.append(f'- {qk}')
        if a['title_cells'][:3]:
            lines += ['', 'Notable headers:', '']
            for t in a['title_cells'][:5]:
                lines.append(f'- {t}')
        lines += ['', '---', '']

    lines += [
        '## Cross-template differences (app vs Excel deviation risk)',
        '',
        '### Peak volume methodology by template group',
        '',
        '| Group | Templates | Excel method | App behavior | Deviation risk |',
        '|-------|-----------|--------------|--------------|----------------|',
        '| **TMR period** | TMR | `period × 60%/15% × 85% / lanes` | `buildTmrPeriod` applies 60%/15% **without** 85% (`app.js` notes removed K-factor) | **HIGH (~18%)** — app hourly ~17.6% higher than Excel |',
        '| **AADT 15%** | AADT-only, Ipswich, GCCC*, Toowoomba* | `AADT × 15% × 85%` with HV split | `DHV_K_FACTOR = 0.85` when `usesCouncilKFactor(site)` — applied for Ipswich/GCCC/Toowoomba, not TMR/Brisbane/Logan | **LOW–MED** for council sources; verify GCCC count path |',
        '| **AADT 12%** | 5-minute counts | Label "85% of 12% of AADT"; hourly from `(100×count/42)×85%` | App may use different peak % and 5-min scaling (`5/60` not `100/42`) | **MED** on scaling factor |',
        '| **Measured counts** | Brisbane, Logan | Peak hour from count profile | App uses measured hourly when `MINIMUM_NONZERO_HOURS` met | **LOW** if same count inputs |',
        '',
        '\\* GCCC / Toowoomba use counts primarily; 15%×85% formulas are fallback/design-check rows.',
        '',
        '### Queue length — universal across all templates',
        '',
        'All eight templates share the same LV/HV multiplier table (based on 6 m LV queued-car spacing):',
        '',
        '| Duration | LV | HV |',
        '|----------|----|----|',
        '| 2 min | 2.4 | 8 |',
        '| 5 min | 6 | 20 |',
        '| 10 min | 12 | 40 |',
        '| 15 min | 18 | 60 |',
        '| 30 min | 36 | 120 |',
        '',
        '**App differences (`calc/tia-calc.js`):**',
        '',
        '1. **LV spacing is speed-dependent:** 6.0 m (posted ≤ 60 km/h), 7.0 m (> 60 km/h) — Excel always uses 6 m equivalent (mult 6 at 5 min).',
        '2. **Model type:** App uses net-overflow / Akçelik-style deterministic queue; Excel uses fixed multiplier on **5-min volumes** derived from hourly.',
        '3. **RT (road train):** App adds RT spacing (63 m per 5 min at base); Excel templates show no RT column in queue factor table.',
        '4. **Merge contingency:** App applies ×1.5 (within taper) / ×2.0 (beyond merge) for SLC scenarios; not evident in standard Excel queue tables.',
        '',
        '### 0.85 K-factor usage',
        '',
        '| Template | Uses 0.85? |',
        '|----------|------------|',
        '| TMR | **Yes** — on all period hourly calcs |',
        '| AADT-only, Ipswich | **Yes** — `AADT×15%×85%` |',
        '| GCCC, Toowoomba | **Yes** — fallback/design rows |',
        '| 5-minute counts | **Yes** — on scaled hourly |',
        '| Brisbane, Logan | **No** — measured counts |',
        '',
        'App: `DHV_K_FACTOR = 0.85` for council AADT paths; explicitly **disabled** for TMR-style hourly (`usesCouncilKFactor` returns false for TMR/Brisbane/Logan). This is **inverted vs Excel TMR** which still applies 85%.',
        '',
        '### VCR / capacity',
        '',
        '- **Excel:** Directional hourly demand (after peak/K factors) divided by Austroads design capacity (DV) per lane; SLC often scales demand by `lanes/open_lanes`.',
        '- **App:** `calculateVCR(hourlyPerLane, adjustedCapacity)` with Austroads terrain/HV/width adjustments in `getAustroadsCapacityProfile`.',
        '- **Risk:** Capacity adjustment stack in app may differ from static DV cell in Excel templates.',
        '',
        '### Recommended regression checks',
        '',
        '1. TMR site AM period: compare Excel `ROUNDUP(period×60%×85%/lanes)` vs app `buildTmrPeriod` (no 85%).',
        '2. Ipswich AADT-only: `AADT×15%×85%` with HV split vs app council K-factor path.',
        '3. 5-min template: `(100/42)×85%` scaling vs app 5-min-to-hourly conversion.',
        '4. Posted speed > 60 km/h: queue at 2/5/10/15 min vs Excel fixed multipliers.',
        '5. SLC + merge: app merge contingency vs Excel (if SLC sheet present in project-specific copies).',
        '',
    ]
    return '\n'.join(lines) + '\n'


def main():
    audits = []
    for fname in sorted(os.listdir(FOLDER)):
        if not fname.endswith(('.xlsx', '.xlsm')):
            continue
        path = os.path.join(FOLDER, fname)
        info = scan_workbook(path)
        peak = infer_peak_formula(info, fname)
        k, k_samples = uses_k_factor(info)
        vcr_notes, vcr_samples = infer_vcr(info)
        hourly_notes, hourly_samples = infer_hourly(info, fname)
        merged_q, queue_full = queue_summary(info)

        audits.append({
            'file': fname,
            'council': council_from_filename(fname),
            'sheets': info['sheets'],
            'peak_formulas': peak,
            'peak_primary': peak[0],
            'peak_samples': (info['formulas_peak'] + info['formulas_85'])[:6],
            'queue_factors': merged_q,
            'queue_full': queue_full,
            'vcr_notes': vcr_notes,
            'vcr_primary': vcr_notes[0],
            'vcr_samples': vcr_samples,
            'k_factor': k,
            'k_samples': k_samples,
            'hourly_notes': hourly_notes,
            'hourly_primary': hourly_notes[0],
            'hourly_samples': hourly_samples,
            'quirks': quirks(fname, info),
            'title_cells': info['title_cells'],
        })

    md = render_md(audits)
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(md)
    print(f'Wrote {OUT}')
    for a in audits:
        print(f"\n{a['file']}: peak={a['peak_primary']}, k={a['k_factor']}, queue={a['queue_full'][:60]}...")


if __name__ == '__main__':
    main()
