"""
Enriches QLD GeoJSON files with heavy vehicle percentage data from the
TMR AADT 2015-2025 XLSX (data.qld.gov.au).

Usage:
    python enrich_qld_hv.py [--download]

    --download  Re-download the XLSX from data.qld.gov.au into %TEMP%
                (omit to reuse cached copy)

Source:
    https://www.data.qld.gov.au/dataset/traffic-census.../download/aadt_2015_2025.xlsx
    Column PC_CLASS_0B = % heavy vehicles (Austroads category 0B)
    Joins to GeoJSON by SITE_ID (integer match)
    For council files without SITE_ID: nearest TMR station within 500 m
"""

import json
import math
import os
import sys
import urllib.request

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

XLSX_URL = (
    "https://www.data.qld.gov.au/dataset/5d74e022-a302-4f40-a594-f1840c92f671"
    "/resource/b856deab-ab20-48f1-85c6-2715e3a1d42c/download/aadt_2015_2025.xlsx"
)
XLSX_CACHE = os.path.join(os.environ.get("TEMP", "/tmp"), "qld_aadt.xlsx")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
QLD_DIR = os.path.join(SCRIPT_DIR, "datasets", "QLD")

# Files with SITE_ID field (integer, matches XLSX SITE_ID)
SITE_ID_FILES = [
    os.path.join(QLD_DIR, "tmr.geojson"),
    os.path.join(QLD_DIR, "toowoomba.geojson"),  # has SITE_ID but zero coordinates
]

# Council files without matching SITE_ID - fall back to nearest-station lookup
NEARBY_FILES = [
    os.path.join(QLD_DIR, "brisbane.geojson"),
    os.path.join(QLD_DIR, "goldcoast.geojson"),
    os.path.join(QLD_DIR, "ipswich.geojson"),
    os.path.join(QLD_DIR, "logan.geojson"),
    os.path.join(QLD_DIR, "tewantin.geojson"),
]

NEARBY_RADIUS_KM = 0.5

ROAD_CLASS_RANK = {
    "freeway": 4,
    "rural_highway": 3,
    "arterial": 2,
    "sub_arterial": 1,
    "local": 0,
}
ROAD_CLASS_MAX_DIFF = 1  # accept same class or one step apart

import re as _re

def infer_road_class(road_name):
    """Mirror of inferRoadCapacityClass() in app.js. Returns class key or None."""
    n = (road_name or "").lower()
    if any(tok in n for tok in ("motorway", "freeway", "mwy", "fwy")):
        return "freeway"
    if _re.search(r"\bm\d+\b", n):
        return "freeway"
    if any(tok in n for tok in ("highway", "hwy", "state route", "national route")):
        return "rural_highway"
    if any(tok in n for tok in ("multi-lane", "divided", "arterial", "state road")):
        return "arterial"
    if any(tok in n for tok in ("sub-arterial", "collector", "boulevard", "blvd", "avenue", " ave")):
        return "sub_arterial"
    if any(tok in n for tok in (" street", " st ", " close", " court", " place", " crescent", " drive", " dr ", " road", " rd ")):
        return "local"
    if n.endswith((" st", " ave", " dr", " rd", " ct", " pl", " cr", " cres", " cl")):
        return "local"
    return None


def class_compatible(council_cls, tmr_cls):
    """True if a TMR station of tmr_cls is acceptable for a council point of council_cls."""
    if council_cls is None or tmr_cls is None:
        return True
    r1 = ROAD_CLASS_RANK.get(council_cls)
    r2 = ROAD_CLASS_RANK.get(tmr_cls)
    if r1 is None or r2 is None:
        return True
    return abs(r1 - r2) <= ROAD_CLASS_MAX_DIFF


def download_xlsx():
    print("Downloading QLD AADT XLSX from data.qld.gov.au...")
    # Download to a temp file with a timeout, then atomically rename. This
    # prevents a network failure or HTTP error from leaving a 0-byte / partial
    # cache that later fails opaquely inside openpyxl (BadZipFile).
    tmp_path = f"{XLSX_CACHE}.tmp"
    req = urllib.request.Request(XLSX_URL, headers={"User-Agent": "tia-dataset-builder"})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp, open(tmp_path, "wb") as out:
            out.write(resp.read())
        os.replace(tmp_path, XLSX_CACHE)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    print(f"  Saved ({os.path.getsize(XLSX_CACHE):,} bytes)")


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def load_hv_data():
    """Returns (site_hv dict, station_xy list)."""
    if not os.path.exists(XLSX_CACHE):
        download_xlsx()

    print(f"Loading XLSX...")
    wb = openpyxl.load_workbook(XLSX_CACHE, read_only=True, data_only=True)
    ws = wb["aadt_2015_2025"]

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {name: i for i, name in enumerate(header)}

    # {site_id: (year, hv_pct, rt_pct, lat, lon, road_name, road_cls)}
    best = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["TRAVEL_DIRECTION"]] or "").strip() != "BOTH DIRECTIONS":
            continue
        try:
            site_id = int(row[idx["SITE_ID"]])
            year = int(row[idx["TRAFFIC_YEAR"]])
            hv_pct = float(row[idx["PC_CLASS_0B"]] or 0)
            rt_pct = sum(float(row[idx[c]] or 0) for c in ["PC_CLASS_2H", "PC_CLASS_2I", "PC_CLASS_2J", "PC_CLASS_2K"] if c in idx)
            lat = float(row[idx["SITE_LATITUDE"]])
            lon = float(row[idx["SITE_LONGITUDE"]])
            road_name = str(row[idx["ROAD_NAME"]] or "").strip()
        except (TypeError, ValueError):
            continue
        if not (0 <= hv_pct <= 100):
            continue
        road_cls = infer_road_class(road_name)
        if site_id not in best or year > best[site_id][0]:
            best[site_id] = (year, hv_pct, rt_pct, lat, lon, road_name, road_cls)

    site_hv = {sid: (v[0], v[1], v[2]) for sid, v in best.items()}
    # station_xy: [(lat, lon, hv_pct, rt_pct, year, road_name, road_cls), ...]
    station_xy = [(v[3], v[4], v[1], v[2], v[0], v[5], v[6]) for v in best.values() if v[3] and v[4]]
    print(f"  {len(site_hv)} unique QLD sites with HV%")
    return site_hv, station_xy


def nearest_hv(lat, lon, station_xy, council_cls=None):
    """Class-aware nearest-station lookup.

    Only matches TMR stations whose inferred road class is compatible with the
    council site's road class (within ROAD_CLASS_MAX_DIFF). Returns the
    matched station's HV%, RT%, year, road name, and road class -- or all
    Nones when no compatible station lies inside the radius.
    """
    best_dist = NEARBY_RADIUS_KM + 1
    best_hv = best_rt = best_year = None
    best_road = best_cls = None
    for slat, slon, hv, rt, year, road_name, road_cls in station_xy:
        if not class_compatible(council_cls, road_cls):
            continue
        d = haversine_km(lat, lon, slat, slon)
        if d < best_dist:
            best_dist = d
            best_hv = hv
            best_rt = rt
            best_year = year
            best_road = road_name
            best_cls = road_cls
    return best_hv, best_rt, best_year, best_road, best_cls


def enrich_by_site_id(path, site_hv):
    if not os.path.exists(path):
        print(f"  Skip (not found): {os.path.basename(path)}")
        return 0, 0
    print(f"Enriching {os.path.basename(path)} by SITE_ID...")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    features = data.get("features", [])
    enriched = 0
    for feat in features:
        if not isinstance(feat, dict):
            continue
        props = feat.get("properties") or {}
        try:
            sid = int(props.get("SITE_ID"))
        except (TypeError, ValueError):
            continue
        if sid in site_hv:
            year, pct, rt = site_hv[sid]
            props["heavy_vehicle_pct"] = round(pct, 1)
            props["heavy_vehicle_year"] = year
            props["hv_source"] = "tmr_site_id"
            if rt > 0:
                props["rt_pct"] = round(rt, 2)
                props["rt_year"] = year
            enriched += 1
    print(f"  {enriched}/{len(features)} features enriched")
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    os.replace(tmp_path, path)  # atomic — never leaves a partial file on failure
    return len(features), enriched


def enrich_by_proximity(path, station_xy):
    if not os.path.exists(path):
        print(f"  Skip (not found): {os.path.basename(path)}")
        return 0, 0
    print(f"Enriching {os.path.basename(path)} by proximity...")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    # Support both FeatureCollection {"features": [...]} and plain list [...]
    is_list = isinstance(data, list)
    features = data if is_list else data.get("features", [])

    enriched = 0
    for feat in features:
        if not isinstance(feat, dict):
            continue
        # Plain-list files store props directly on the object
        props = feat if is_list else (feat.get("properties") or {})
        if not isinstance(props, dict):
            continue
        # Re-enrich every run to honour latest class-compatibility rules.
        # (Previous proximity-derived values get overwritten.)
        geom = {} if is_list else (feat.get("geometry") or {})
        coords = geom.get("coordinates") if isinstance(geom, dict) else None
        lat = lon = None
        if isinstance(coords, (list, tuple)) and len(coords) >= 2:
            try:
                lon = float(coords[0])
                lat = float(coords[1])
            except (TypeError, ValueError):
                lat = lon = None
        if lat is None or lon is None:
            lat = props.get("LATITUDE") or props.get("latitude") or props.get("wgs84_latitude") or props.get("Latitude")
            lon = props.get("LONGITUDE") or props.get("longitude") or props.get("wgs84_longitude") or props.get("Longitude")
        try:
            lat, lon = float(lat), float(lon)
        except (TypeError, ValueError):
            continue
        if lat == 0.0 and lon == 0.0:
            continue
        # Classify the council survey road from its road name.
        road_name = props.get("ROAD_NAME") or props.get("road_name") or ""
        council_cls = infer_road_class(road_name)
        hv, rt, hv_year, src_road, src_cls = nearest_hv(lat, lon, station_xy, council_cls)
        if hv is not None:
            props["heavy_vehicle_pct"] = round(hv, 1)
            props["heavy_vehicle_year"] = hv_year
            props["rt_pct"] = round(rt, 2) if rt is not None else 0.0
            props["rt_year"] = hv_year if (rt is not None and rt > 0) else None
            props["hv_source"] = "tmr_proximity"
            props["hv_source_road_name"] = src_road or ""
            props["hv_source_road_class"] = src_cls or ""
            enriched += 1
        else:
            # No compatible match -- strip any previous proximity-derived values
            # so the UI can fall back to smart defaults.
            for k in ("heavy_vehicle_pct", "heavy_vehicle_year", "rt_pct", "rt_year",
                     "hv_source", "hv_source_road_name", "hv_source_road_class"):
                props.pop(k, None)

    print(f"  {enriched}/{len(features)} features enriched")
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    os.replace(tmp_path, path)  # atomic — never leaves a partial file on failure
    return len(features), enriched


def main():
    if "--download" in sys.argv and os.path.exists(XLSX_CACHE):
        os.remove(XLSX_CACHE)

    site_hv, station_xy = load_hv_data()

    total = 0
    for path in SITE_ID_FILES:
        _, n = enrich_by_site_id(path, site_hv)
        total += n
    for path in NEARBY_FILES:
        _, n = enrich_by_proximity(path, station_xy)
        total += n

    print(f"\nDone. {total} QLD features enriched with heavy_vehicle_pct.")
    print("Next: run upload_enriched.py to compress and push to Firebase Storage.")


if __name__ == "__main__":
    main()
