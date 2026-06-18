#!/usr/bin/env python3
"""
step1_build_tmr_profiles_2024.py
-------------------------------
Downloads the 2024 TMR hourly traffic data from QLD Open Data and rebuilds
tmr_profiles.pkl with:

  ? Statewide AADT band profiles   (same structure as before, but 2024 data)
  ? LGA-specific band profiles     (NEW -- per-council hourly shapes)
  ? Site-level profiles            (for Brisbane spatial matching)
  ? Brisbane-region fallback       (unchanged purpose)

The LGA profiles let every council use TMR stations *within their own
boundaries* instead of the statewide average -- more accurate for Gold Coast
tourist roads, Toowoomba inland patterns, etc.

USAGE
-----
  python scripts/step1_build_tmr_profiles_2024.py

OUTPUT
------
  tmr_profiles.pkl          -- updated profiles (old file backed up)

REQUIRES
--------
  pip install openpyxl requests
"""

from __future__ import annotations

import io
import math
import os
import pickle
import shutil
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

# -- Paths ---------------------------------------------------------------------
ROOT        = Path(__file__).resolve().parent.parent
PKL_PATH    = ROOT / "tmr_profiles.pkl"
CACHE_DIR   = ROOT / "datasets" / "QLD" / "2024_update" / "source_data"

# -- Source URLs ---------------------------------------------------------------
HOURLY_URL = (
    "https://www.data.qld.gov.au/dataset/5334361b-3d7b-476d-9776-04dcd4a2d388"
    "/resource/ef3b4602-696a-4a8b-9934-c1f76e147834"
    "/download/qld-traffic-data-average-by-hour-by-day-2024.xlsx"
)
AADT_URL = (
    "https://www.data.qld.gov.au/dataset/5d74e022-a302-4f40-a594-f1840c92f671"
    "/resource/b856deab-ab20-48f1-85c6-2715e3a1d42c"
    "/download/aadt_2014_2024.xlsx"
)

# -- Config --------------------------------------------------------------------
AADT_BANDS      = [(0,5_000),(5_000,15_000),(15_000,30_000),(30_000,60_000),(60_000,10**9)]
BRISBANE_BBOX   = {"lat_min":-27.8,"lat_max":-27.1,"lon_min":152.6,"lon_max":153.5}

# Councils we want LGA-specific profiles for (name must match LGA_NAME in AADT census)
LGA_TARGETS = {
    "Brisbane City":            "brisbane",
    "Gold Coast City":          "goldcoast",
    "Logan City":               "logan",
    "Ipswich City":             "ipswich",
    "Toowoomba Regional":       "toowoomba",
    "Moreton Bay City":         "moreton_bay",
    "Sunshine Coast Regional":  "sunshine_coast",
    "Redland City":             "redland",
}

HEADERS = {"User-Agent": "Mozilla/5.0 (Crompton-TIA-Updater/1.0)"}


# -- Helpers -------------------------------------------------------------------

def _dl(url: str, cache_name: str) -> bytes:
    """Download with cache."""
    CACHE_DIR.mkdir(exist_ok=True)
    cache_file = CACHE_DIR / cache_name
    if cache_file.exists():
        print(f"    (cached) {cache_name}")
        return cache_file.read_bytes()
    print(f"    Downloading {cache_name} ...", end=" ", flush=True)
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=120) as r:
        data = r.read()
    cache_file.write_bytes(data)
    print(f"{len(data)/1024/1024:.1f} MB")
    return data


def _normalise(lst: list[float]) -> list[float]:
    s = sum(lst)
    return [v / s for v in lst] if s > 0 else [1/24] * 24


def get_band(aadt: float) -> tuple[int,int]:
    for lo, hi in AADT_BANDS:
        if lo <= aadt < hi:
            return (lo, hi)
    return AADT_BANDS[-1]


# -- Load AADT census -> LGA lookup ---------------------------------------------

def load_lga_lookup(aadt_xlsx: bytes) -> dict[int, str]:
    """Return {site_id: lga_name} using 2024 rows from the AADT census."""
    print("  Parsing AADT census for LGA assignments ...", end=" ", flush=True)
    wb  = openpyxl.load_workbook(io.BytesIO(aadt_xlsx), read_only=True, data_only=True)
    ws  = wb["aadt_2014_2024"]
    lga = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != 2024:   # TRAFFIC_YEAR column
            continue
        site_id  = row[4]    # SITE_ID
        lga_name = row[22]   # LGA_NAME
        if site_id and lga_name:
            lga[int(site_id)] = lga_name
    wb.close()
    print(f"{len(lga):,} sites")
    return lga


# -- Parse 2024 hourly sheet ---------------------------------------------------

def parse_hourly(hourly_xlsx: bytes, lga_lookup: dict[int,str]) -> dict:
    """
    Returns:
        site_profiles  : {site_id: {lat, lon, wd_pct, we_pct, aadt_wd, lga}}
        band_profiles  : {(lo,hi): {wd, we, n}}
        lga_band_profiles : {lga_key: {(lo,hi): {wd, we, n}}}
        brisbane_fallback : {wd_pct, we_pct, n_sites}
    """
    print("  Parsing 2024 hourly data ...", end=" ", flush=True)
    wb = openpyxl.load_workbook(io.BytesIO(hourly_xlsx), read_only=True, data_only=True)
    ws = wb.active

    # Accumulate per-site aggregated weekday/weekend profiles
    # Structure: site_id -> {lat, lon, dirs: {dir -> [24]}, we_dirs: {dir -> [24]}}
    raw: dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        site_id = row[0]   # SITE_ID
        if site_id is None:
            continue
        site_id  = int(site_id)
        lon      = row[3]   # LONGITUDE
        lat      = row[4]   # LATITUDE
        gazettal = row[9]   # GAZETTAL_DIRECTION
        hours    = row[10]  # HOURS "X to Y"
        wd_avg   = row[18]  # WEEKDAY_AVERAGE
        we_avg   = row[19]  # WEEKEND_AVERAGE
        try:
            hr = int(str(hours).split(" to ")[0])
        except Exception:
            continue

        if site_id not in raw:
            raw[site_id] = {"lat": lat, "lon": lon, "wd": [0]*24, "we": [0]*24}
        raw[site_id]["lat"] = lat or raw[site_id]["lat"]
        raw[site_id]["lon"] = lon or raw[site_id]["lon"]
        raw[site_id]["wd"][hr] += (wd_avg or 0)
        raw[site_id]["we"][hr] += (we_avg or 0)

    wb.close()
    print(f"{len(raw):,} sites")

    # Build normalised profiles
    site_profiles: dict = {}
    # accumulators: band -> list of wd/we profiles
    band_wd: dict[tuple,list] = defaultdict(list)
    band_we: dict[tuple,list] = defaultdict(list)
    # LGA accumulators: lga_key -> band -> list of profiles
    lga_band_wd: dict[str,dict[tuple,list]] = defaultdict(lambda: defaultdict(list))
    lga_band_we: dict[str,dict[tuple,list]] = defaultdict(lambda: defaultdict(list))

    for sid, sd in raw.items():
        aadt_wd = sum(sd["wd"])
        if aadt_wd == 0:
            continue
        wp = _normalise(sd["wd"])
        ep = _normalise(sd["we"])
        b  = get_band(aadt_wd)
        lga_name = lga_lookup.get(sid, "")
        lga_key  = LGA_TARGETS.get(lga_name, "")

        site_profiles[sid] = {
            "lat":     sd["lat"],
            "lon":     sd["lon"],
            "wd_pct":  wp,
            "we_pct":  ep,
            "aadt_wd": int(aadt_wd),
            "lga":     lga_key,
        }
        band_wd[b].append(wp)
        band_we[b].append(ep)
        if lga_key:
            lga_band_wd[lga_key][b].append(wp)
            lga_band_we[lga_key][b].append(ep)

    # Statewide band profiles
    band_profiles: dict = {}
    for b in AADT_BANDS:
        if b not in band_wd:
            continue
        n = len(band_wd[b])
        avg_wd = [sum(p[h] for p in band_wd[b]) / n for h in range(24)]
        avg_we = [sum(p[h] for p in band_we[b]) / n for h in range(24)]
        band_profiles[b] = {"wd": _normalise(avg_wd), "we": _normalise(avg_we), "n": n}

    # LGA-specific band profiles
    lga_band_profiles: dict = {}
    for lga_key, bmap in lga_band_wd.items():
        lga_band_profiles[lga_key] = {}
        for b, profiles in bmap.items():
            n = len(profiles)
            avg_wd = [sum(p[h] for p in profiles) / n for h in range(24)]
            avg_we = [sum(p[h] for p in lga_band_we[lga_key][b]) / n for h in range(24)]
            lga_band_profiles[lga_key][b] = {"wd": _normalise(avg_wd), "we": _normalise(avg_we), "n": n}

    # For bands missing in a given LGA, fill with nearest statewide band
    for lga_key in lga_band_profiles:
        for b in AADT_BANDS:
            if b not in lga_band_profiles[lga_key] and b in band_profiles:
                lga_band_profiles[lga_key][b] = band_profiles[b]
                lga_band_profiles[lga_key][b]["filled_from_statewide"] = True

    # Brisbane region fallback
    bb = BRISBANE_BBOX
    bw = [0.0]*24; be = [0.0]*24; bn = 0
    for sp in site_profiles.values():
        if (sp["lat"] and sp["lon"] and
                bb["lat_min"] <= sp["lat"] <= bb["lat_max"] and
                bb["lon_min"] <= sp["lon"] <= bb["lon_max"]):
            for h in range(24):
                bw[h] += sp["wd_pct"][h]
                be[h] += sp["we_pct"][h]
            bn += 1
    brisbane_fallback = {"wd_pct": _normalise(bw), "we_pct": _normalise(be), "n_sites": bn}

    return site_profiles, band_profiles, lga_band_profiles, brisbane_fallback


# -- Print summary -------------------------------------------------------------

def _peak(pcts: list[float]) -> str:
    ph = max(range(24), key=lambda h: pcts[h])
    return f"{ph:02d}:00 ({pcts[ph]*100:.1f}%)"


def print_summary(band_profiles, lga_band_profiles, brisbane_fallback):
    print("\n  === Statewide band profiles ===")
    for b, p in sorted(band_profiles.items()):
        print(f"    {b[0]:>6}-{b[1]:<12,}  n={p['n']:>4}  "
              f"WD peak={_peak(p['wd'])}  night={sum(p['wd'][:5])*100:.1f}%")

    print("\n  === LGA-specific profiles ===")
    for lga_key in sorted(lga_band_profiles):
        lga_bps = lga_band_profiles[lga_key]
        mid_band = (15_000, 30_000)
        bp = lga_bps.get(mid_band, lga_bps[sorted(lga_bps)[0]])
        statewide = " (statewide)" if bp.get("filled_from_statewide") else ""
        print(f"    {lga_key:<18}  {len(lga_bps)} bands  "
              f"WD peak={_peak(bp['wd'])}{statewide}")
    print(f"\n  Brisbane fallback: {brisbane_fallback['n_sites']} sites")


# -- Main ----------------------------------------------------------------------

def main():
    print("\n[1/4] Downloading data files ...")
    hourly_xlsx = _dl(HOURLY_URL, "tmr_2024_hourly.xlsx")
    aadt_xlsx   = _dl(AADT_URL,   "tmr_2024_aadt.xlsx")

    print("[2/4] Loading LGA assignments ...")
    lga_lookup = load_lga_lookup(aadt_xlsx)

    print("[3/4] Building profiles ...")
    site_profiles, band_profiles, lga_band_profiles, bris_fallback = \
        parse_hourly(hourly_xlsx, lga_lookup)

    print_summary(band_profiles, lga_band_profiles, bris_fallback)

    print("\n[4/4] Saving tmr_profiles.pkl ...")
    # Backup existing file
    if PKL_PATH.exists():
        bak = PKL_PATH.with_suffix(".pkl.bak")
        shutil.copy2(PKL_PATH, bak)
        print(f"  Backed up existing -> {bak.name}")

    payload = {
        "band_profiles":     band_profiles,
        "site_profiles":     site_profiles,
        "brisbane_fallback": bris_fallback,
        "lga_band_profiles": lga_band_profiles,
        "data_year":         2024,
        "lga_targets":       LGA_TARGETS,
    }
    with open(PKL_PATH, "wb") as f:
        pickle.dump(payload, f)

    size_kb = PKL_PATH.stat().st_size / 1024
    print(f"  Saved {PKL_PATH.name}  ({size_kb:.0f} KB)")
    print(f"  Sites: {len(site_profiles):,}  |  Statewide bands: {len(band_profiles)}"
          f"  |  LGA profiles: {len(lga_band_profiles)}")
    print("\nDone. Run step2_download_council_counts.py next.\n")


if __name__ == "__main__":
    main()
